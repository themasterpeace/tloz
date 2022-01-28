#from django.http import response
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, ListView,CreateView,UpdateView,DeleteView
from .forms import *
from .models import *
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
# Create your views here.

@permission_required('zelda.add_registro')
def control(request):
    data ={
        'form': RegistroForm()
    }
    
    if request.method == 'POST':
        formulario= RegistroForm(data=request.POST, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Registro Guardado correctamente")
        else:
            data["form"] = formulario
            
    return render(request, 'zelda/home.html', data)

@permission_required('zelda.view_registro')
def listar(request):
    
    registros = Registro.objects.all()
    
    data ={
        'registros': registros
    }
    return render(request, 'zelda/listar.html', data)

@permission_required('zelda.change_registro')
def modificar(request, id):
    
    registro = get_object_or_404(Registro, id=id)
    
    data={
        'form': RegistroForm(instance=registro)
    }
    
    if request.method == 'POST':
        formulario = RegistroForm(data=request.POST, instance=registro, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Registro modificado correctamente")
            return redirect(to="listar")
        data["form"] = formulario
    
    return render(request, 'zelda/modificar.html', data)

@permission_required('zelda.delete_registro')
def eliminar(request, id):
    producto=get_object_or_404(Registro, id=id)
    producto.delete()
    messages.success(request, "Registro Eliminado Exitosamente")
    return redirect(to="listar")


class ReporteRegistrosExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        registros = Registro.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE REGISTROS'
        
        ws.merge_cells('B1:R1')
        
        ws['A3']='Fecha llenado'
        ws['B3']='Hora llenado'
        ws['C3']='Placa'
        ws['D3']='Piloto'
        ws['E3']='Ruta'
        ws['F3']='Serie factura'
        ws['G3']='No. Factura'
        ws['H3']='Total Llenado'
        ws['I3']='Galones Llenados'
        ws['J3']='Tipo Combustible'
        ws['K3']='Bomba Llenado'
        ws['L3']='Precio Por Galon'
        ws['M3']='Kilometraje Inicial'
        ws['N3']='Kilometraje Final'
        ws['O3']='Recorrido'
        ws['P3']='Redimiento X Galon'
        ws['Q3']='Manifiesto'
        ws['R3']='Observaciones'
               
        cont = 4
        
        for registro in registros:
            ws.cell(row = cont, column=1).value = registro.fecha
            ws.cell(row = cont, column=2).value = registro.hora_llenado
            ws.cell(row = cont, column=3).value = registro.placa
            ws.cell(row = cont, column=4).value = registro.piloto
            ws.cell(row = cont, column=5).value = registro.ruta
            ws.cell(row = cont, column=6).value = registro.serie
            ws.cell(row = cont, column=7).value = registro.no_factura
            ws.cell(row = cont, column=8).value = registro.total
            ws.cell(row = cont, column=9).value = registro.galones
            ws.cell(row = cont, column=10).value = registro.tipo_combustible
            ws.cell(row = cont, column=11).value = registro.bomnba
            ws.cell(row = cont, column=12).value = registro.precioxgalon
            ws.cell(row = cont, column=13).value = registro.kminicial
            ws.cell(row = cont, column=14).value = registro.kmfinal
            ws.cell(row = cont, column=15).value = registro.recorrido
            ws.cell(row = cont, column=16).value = registro.rendxgalon
            ws.cell(row = cont, column=17).value = registro.manifiesto
            ws.cell(row = cont, column=18).value = registro.observaciones
            cont+=1
            
        nombre_archivo = "ReporteRegistrosExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']= content
        wb.save(response)
        return response
    


            
        
        